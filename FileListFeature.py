import os
import sys
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from Feature import Feature
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


class FileListFeature(Feature):
    def run(self, frame, return_callback):
        for widget in frame.winfo_children():
            widget.destroy()

        label = ttk.Label(frame, text="Dateien in Ordnern auflisten + Excel-Export", font=("Arial", 12))
        label.pack(pady=10)

        table_frame = ttk.Frame(frame)
        table_frame.pack(fill="both", expand=True)

        tree = ttk.Treeview(table_frame, columns=("ordner", "datei", "pfad", "link"), show="headings")
        tree.heading("ordner", text="Ordnername")
        tree.heading("datei", text="Dateiname")
        tree.heading("pfad", text="Pfad")
        tree.heading("link", text="Link")
        for col in ("ordner", "datei", "pfad", "link"):
            tree.column(col, width=150 if col != "pfad" else 300, anchor="w")
        tree.pack(fill="both", expand=True)

        file_data = []

        def on_double_click(event):
            item_id = tree.identify_row(event.y)
            if item_id:
                values = tree.item(item_id, "values")
                pfad = values[2]  # 3. Spalte = Dateipfad
                try:
                    if os.path.exists(pfad):
                        if os.name == 'nt':
                            os.startfile(pfad)
                        elif os.name == 'posix':
                            import subprocess
                            subprocess.run(["open" if sys.platform == "darwin" else "xdg-open", pfad])
                except Exception as e:
                    messagebox.showerror("Fehler", f"Konnte Datei nicht öffnen:\n{pfad}\n{e}")

        tree.bind("<Double-1>", on_double_click)

        def list_files():
            folder = filedialog.askdirectory(title="Wähle einen Ordner")
            if not folder:
                return

            tree.delete(*tree.get_children())
            file_data.clear()

            for root, _, files in os.walk(folder):
                ordnername = os.path.basename(root)
                for file in sorted(files):  # alphabetisch sortiert
                    pfad = os.path.join(root, file)
                    file_data.append((ordnername, file, pfad, "Öffnen"))

            # nach Ordnernamen sortieren
            file_data.sort(key=lambda x: x[0])

            for row in file_data:
                tree.insert("", "end", values=row)

        def export_to_excel():
            if not file_data:
                messagebox.showwarning("Keine Daten", "Bitte zuerst einen Ordner auswählen.")
                return

            filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel-Dateien", "*.xlsx")])
            if not filepath:
                return

            wb = Workbook()
            ws = wb.active
            ws.title = "Dateiliste"

            headers = ["Ordnername", "Dateiname", "Pfad", "Link"]
            ws.append(headers)

            bold_font = Font(bold=True)
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = bold_font

            for row in file_data:
                ordner, datei, pfad, _ = row
                ws.append([ordner, datei, pfad, "Öffnen"])

            for row_idx in range(2, len(file_data) + 2):
                cell = ws.cell(row=row_idx, column=4)
                pfad = ws.cell(row=row_idx, column=3).value
                cell.hyperlink = f"file://{pfad}"
                cell.style = "Hyperlink"

            for col_idx, col in enumerate(headers, 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 40 if col == "Pfad" else 25

            wb.save(filepath)
            messagebox.showinfo("Export abgeschlossen", f"Excel-Datei gespeichert:\n{filepath}")

        ttk.Button(frame, text="Ordner auswählen & Dateien anzeigen", command=list_files).pack(pady=5)
        ttk.Button(frame, text="Excel exportieren", command=export_to_excel).pack(pady=5)
        ttk.Button(frame, text="Zurück", command=return_callback).pack(pady=10)